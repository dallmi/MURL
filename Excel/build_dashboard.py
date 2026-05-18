"""Generate the MURL Excel Dashboard.

Native-Excel approach — uses Excel Table auto-filter (built-in search box,
contains-search, multi-select) instead of dynamic-array FILTER formulas. KPIs
use SUBTOTAL to respect the filter. Compatible with Excel for Mac, Excel for
the Web (OneDrive sharing), and Excel desktop.

Run:
    python build_dashboard.py             # with ~150 demo rows
    python build_dashboard.py --empty     # empty template for production data

Output: ./murl_dashboard.xlsx
"""

import argparse
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT_PATH = Path(__file__).parent / "murl_dashboard.xlsx"

# Corporate palette (with FF alpha prefix for openpyxl)
RED = "FFE60000"
BLACK = "FF000000"
WHITE = "FFFFFFFF"
PAGE_BG = "FFF7F7F5"
PASTEL_I = "FFECEBE4"
GRAY_I = "FFCCCABC"
GRAY_IV = "FF7A7870"
GRAY_VI = "FF404040"
RAG_GREEN = "FF6F7A1A"
RAG_AMBER = "FFE4A911"
LAKE = "FF0C7EC6"

DATA_HEADERS = [
    "Labels", "Reference", "Summary", "Service project", "Status",
    "Requester", "Type", "Region(s)", "Line manager", "MURL Name",
    "Activation", "Properties", "GOTO/MURL Owner 1", "GOTO/MURL Owner 2",
    "Business Division requested for", "Target Default",
]
COL_WIDTHS = [14, 14, 50, 14, 22, 22, 22, 14, 22, 24, 14, 24, 22, 22, 30, 60]

# Status column index in DATA_HEADERS (1-based for SUBTOTAL ref)
STATUS_COL_IDX = DATA_HEADERS.index("Status") + 1


def main():
    parser = argparse.ArgumentParser(description="Build MURL Excel Dashboard")
    parser.add_argument("--empty", action="store_true",
                        help="Generate empty template (default: seed ~150 demo rows)")
    parser.add_argument("--seed-count", type=int, default=150,
                        help="Number of seed rows when seeding (default: 150)")
    args = parser.parse_args()

    wb = Workbook()
    ws = wb.active
    ws.title = "MURL Dashboard"
    ws_help = wb.create_sheet("Anleitung")

    seed_rows = [] if args.empty else generate_seed_rows(args.seed_count)
    build_main_sheet(ws, seed_rows)
    build_help_sheet(ws_help)

    ws.sheet_view.showGridLines = False
    ws_help.sheet_view.showGridLines = False
    wb.active = 0

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")


def build_main_sheet(ws, rows):
    """Single-sheet dashboard: title, KPIs, then Excel Table with auto-filter."""
    last_col = len(DATA_HEADERS)
    last_col_letter = get_column_letter(last_col)

    # Column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title row
    ws.merge_cells(f"A1:{last_col_letter}1")
    title = ws["A1"]
    title.value = "MURL Dashboard"
    title.font = Font(name="Calibri", size=24, bold=True, color=BLACK)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 40
    red_border = Border(bottom=Side(border_style="medium", color=RED))
    for col in range(1, last_col + 1):
        ws.cell(row=1, column=col).border = red_border

    # Subtitle
    ws.merge_cells(f"A2:{last_col_letter}2")
    sub = ws["A2"]
    sub.value = (
        "Klicke auf den Pfeil im Spaltenkopf der Tabelle, um zu filtern — "
        "Suchfeld, Mehrfachauswahl und Contains-Suche sind eingebaut."
    )
    sub.font = Font(name="Calibri", size=11, color=GRAY_IV)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

    # KPI cards (row 4 label, row 5 value). SUBTOTAL respects auto-filter.
    table_status_ref = f"tblData[Status]"
    table_ref_col_ref = f"tblData[Reference]"

    kpis = [
        # Total visible rows
        ("Total", f'=SUBTOTAL(103,{table_ref_col_ref})', RED),
        ("Active", _kpi_count_visible_formula("Active"), RAG_GREEN),
        ("Scheduled", _kpi_count_visible_formula("Scheduled Activation"), LAKE),
        ("Pending", _kpi_count_visible_formula("Pending Change"), RAG_AMBER),
        ("Deactivated", _kpi_count_visible_formula("Deactivated"), GRAY_IV),
    ]
    # 3-col-wide KPI cards
    for i, (label, formula, accent) in enumerate(kpis):
        start = 1 + i * 3
        end = start + 2

        ws.merge_cells(start_row=4, end_row=4, start_column=start, end_column=end)
        lc = ws.cell(row=4, column=start)
        lc.value = label.upper()
        lc.font = Font(name="Calibri", size=10, bold=True, color=GRAY_IV)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        lc.fill = PatternFill("solid", fgColor=PAGE_BG)
        lc.border = Border(
            top=Side(border_style="thick", color=accent),
            left=Side(border_style="thin", color=GRAY_I),
            right=Side(border_style="thin", color=GRAY_I),
        )

        ws.merge_cells(start_row=5, end_row=5, start_column=start, end_column=end)
        vc = ws.cell(row=5, column=start)
        vc.value = formula
        vc.font = Font(name="Calibri", size=22, color=BLACK)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        vc.fill = PatternFill("solid", fgColor=PAGE_BG)
        vc.border = Border(
            left=Side(border_style="thin", color=GRAY_I),
            right=Side(border_style="thin", color=GRAY_I),
            bottom=Side(border_style="thin", color=GRAY_I),
        )
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 44

    # Data table — header row at row 7, data rows from row 8
    header_row = 7
    for col_idx, header in enumerate(DATA_HEADERS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=GRAY_VI)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[header_row].height = 26

    # Data rows — skip empty values to avoid malformed inlineStr cells
    for row_idx, row in enumerate(rows, start=header_row + 1):
        for col_idx, header in enumerate(DATA_HEADERS, 1):
            val = row.get(header)
            if val:
                ws.cell(row=row_idx, column=col_idx, value=val)

    last_data_row = header_row + max(len(rows), 1)
    table_range = f"A{header_row}:{last_col_letter}{last_data_row}"
    tbl = Table(displayName="tblData", ref=table_range)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)

    ws.freeze_panes = f"A{header_row + 1}"


def _kpi_count_visible_formula(status_value):
    """COUNTIFS-style count that respects auto-filter via SUBTOTAL+OFFSET trick.

    Counts rows where Status equals `status_value` AND the row is visible
    (not hidden by Excel's auto-filter on the table).
    """
    return (
        '=SUMPRODUCT('
        'SUBTOTAL(103,OFFSET(tblData[[#Headers],[Status]],'
        'ROW(tblData[Status])-ROW(tblData[[#Headers],[Status]]),0))'
        f'*(tblData[Status]="{status_value}")'
        ')'
    )


def generate_seed_rows(count):
    random.seed(42)
    people = [
        "Demetz Domenik", "Potter Holly", "Garic Ivona", "Oldani Marina",
        "Davey Christopher", "Rodriguez Karen", "Caldelari Sandro", "Kaufmann Maya",
        "Stofer Anja", "Sison Jeymarc", "Ulrich Valentin", "Larado David",
        "Frey Joel", "Curry-Curry Sonja", "Orland Agnese", "Wilkening Janis",
        "Belik Dorota", "Pezzini Adele", "Stanisic Sava", "Spencer Jeffrey",
        "Donnenfeld Jake", "Muchtadi Imanda", "Thielmann Phil", "Finck Susanne",
        "Brigatto Gianluca", "Yeskev Alexander", "Heindel Tilo", "Wallin Mark",
        "Stunde Andrew", "Vandermey Joshua", "Picosrick Michael", "Wilkening Jens",
        "Johnson Patrick", "Maslehuk Caroline", "Schmid Reto", "Keller Andrea",
    ]
    statuses_weighted = (
        ["Active"] * 14
        + ["Scheduled Activation"] * 3
        + ["Pending Change"] * 2
        + ["Deactivated"] * 1
    )
    regions_weighted = ["Switzerland"] * 6 + ["Global"] * 2 + [""] * 2
    divisions = ["GWM", "P&CB", "IB", "GF", "GWM Americas", "AM", ""]
    activations_weighted = ["Immediate"] * 4 + ["Scheduled"] * 1
    properties_pool = [
        "", "", "", "Language-specific", "Technical", "Prioritize",
        "Language-specific;Technical",
    ]
    murl_codes = [
        "newera", "dxsj5", "swissbanking", "dailybanking", "klahr", "young-account",
        "freiburg", "risk-insights", "wallis", "romandie", "valais", "geneve",
        "brdyj", "cvkkw", "ggrop", "lwol7", "fipd", "vbd", "lpd", "fxd",
        "cwz-immobilien", "swiss-hotels", "m3pzs", "fei-dressur", "fei-springen",
        "talents-program", "academic-family", "youth-students", "pension-vested",
        "card-debit", "card-credit", "online-banking", "mobile-banking", "advisor-locator",
        "esg-investing", "crypto-hub", "wealth-planning", "retirement-pillar3",
    ]
    url_segments = [
        "wealth-management/who-we-serve",
        "services/accounts-and-cards/banking-youth-and-students",
        "microsites/regional-impact/lake-zurich",
        "academia-family",
        "online-banking/login",
        "credit-cards/comparison",
        "pension/integrate-vested",
        "services/accounts-and-cards/private-account",
        "wealth-management/global-family-office",
        "investment/crypto-overview",
        "services/risk-information",
        "events/cwz2026-immo",
        "designation-disclosures",
        "communities/leadership/founders",
        "advisory/regional-team/zurich",
    ]
    summary_subdomains = ["www", "secure", "online", "promo"]

    rows = []
    for i in range(count):
        ref_num = 18001 + i
        status = random.choice(statuses_weighted)
        owner1 = random.choice(people)
        owner2 = random.choice(people)
        requester = random.choice(people)
        line_manager = random.choice(people)
        region = random.choice(regions_weighted)
        division = random.choice(divisions)
        activation = random.choice(activations_weighted)
        properties = random.choice(properties_pool)
        murl_name = random.choice(murl_codes)
        url_path = random.choice(url_segments)
        sub = random.choice(summary_subdomains)
        target_lang = random.choice(["en", "de", "fr", "it"])

        target_url = f"https://www.corp.example/ch/{target_lang}/{url_path}.html"
        summary = f"{sub}.corp.example/{url_path}"

        labels = f"ITO-{random.randint(10000, 99999)}" if random.random() < 0.15 else ""

        rows.append({
            "Labels": labels,
            "Reference": f"MURL-{ref_num}",
            "Summary": summary,
            "Service project": "MURL",
            "Status": status,
            "Requester": requester,
            "Type": "Marketing URL (MURL)",
            "Region(s)": region,
            "Line manager": line_manager,
            "MURL Name": murl_name,
            "Activation": activation,
            "Properties": properties,
            "GOTO/MURL Owner 1": owner1,
            "GOTO/MURL Owner 2": owner2,
            "Business Division requested for": division,
            "Target Default": target_url,
        })
    return rows


def build_help_sheet(ws):
    lines = [
        ("Anleitung — MURL Dashboard", 18, True, BLACK),
        ("", 11, False, BLACK),
        ("1. Filtern", 13, True, BLACK),
        ("Klicke auf den kleinen Pfeil rechts in jedem Spaltenkopf der Tabelle.", 11, False, GRAY_VI),
        ("Im Dropdown:", 11, False, GRAY_VI),
        ("   • Suchfeld oben — tippe einen Teilstring, die Werte werden gefiltert (Contains).", 11, False, GRAY_VI),
        ("   • Checkboxen — wähle einen oder mehrere Werte aus.", 11, False, GRAY_VI),
        ("   • 'Textfilter' → 'Enthält' für noch flexiblere Suchen.", 11, False, GRAY_VI),
        ("Mehrere Spalten gleichzeitig filtern: einfach in jeder Spalte den Filter setzen — sie wirken kombiniert (UND).", 11, False, GRAY_VI),
        ("Filter zurücksetzen: im Dropdown 'Filter aufheben' klicken oder Alt+Daten → Löschen.", 11, False, GRAY_VI),
        ("", 11, False, BLACK),
        ("2. KPIs", 13, True, BLACK),
        ("Die 5 KPI-Kacheln oben aktualisieren sich automatisch basierend auf dem aktuellen Filter (sichtbare Zeilen).", 11, False, GRAY_VI),
        ("", 11, False, BLACK),
        ("3. CSV-Daten laden", 13, True, BLACK),
        ("Original-Quelle ist ein CSV-Export aus dem Crosstab-Download des Tableau-Reports.", 11, False, GRAY_VI),
        ("Variante A — Power Query (Excel Desktop, live aktualisierbar):", 11, True, GRAY_VI),
        ("   Daten → Daten abrufen → Aus Text/CSV → Datei wählen → Laden in → Tabelle 'tblData' überschreiben.", 11, False, GRAY_VI),
        ("Variante B — Manuell (auch in Excel for the Web):", 11, True, GRAY_VI),
        ("   CSV in Excel öffnen, alle Datenzeilen markieren (ohne Header), kopieren.", 11, False, GRAY_VI),
        ("   Im Dashboard die zweite Tabellenzeile (direkt unter den Headern) klicken, einfügen.", 11, False, GRAY_VI),
        ("   Die Tabelle erweitert sich automatisch — KPIs aktualisieren sich.", 11, False, GRAY_VI),
        ("", 11, False, BLACK),
        ("4. Teilen via OneDrive", 13, True, BLACK),
        ("Datei in OneDrive ablegen, mit Empfängern teilen — funktioniert in Excel for the Web ohne Add-Ins.", 11, False, GRAY_VI),
        ("", 11, False, BLACK),
        ("5. Spalten-Mapping zum Tableau-Dashboard", 13, True, BLACK),
        ("Reference = ID, GOTO/MURL Owner 1/2 = Owner 1/2, Target Default = Target URL, Requester = Requestor.", 11, False, GRAY_VI),
        ("Tableau-Felder NICHT im CSV enthalten: Is Tiny URL, Created, Expiration.", 11, False, GRAY_VI),
    ]
    for i, (text, size, bold, color) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
        cell.alignment = Alignment(vertical="center", wrap_text=False)
    ws.column_dimensions["A"].width = 130

    ws.cell(row=1, column=1).border = Border(
        bottom=Side(border_style="medium", color=RED)
    )
    ws.row_dimensions[1].height = 32


if __name__ == "__main__":
    main()
