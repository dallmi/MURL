from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.config import XLSX_PATH, XLSX_COLUMNS, OUTPUT_DIR


HEADER_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Arial", size=10)
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=False)
BORDER = Border(
    bottom=Side(style="thin", color="CCCABC"),
    right=Side(style="thin", color="CCCABC"),
)
ALT_ROW_FILL = PatternFill(start_color="F8F7F2", end_color="F8F7F2", fill_type="solid")


def generate_report(records: list[dict]) -> None:
    """Generate XLSX report from transformed records."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "MURL Report"

    headers = [col[0] for col in XLSX_COLUMNS]
    field_keys = [col[1] for col in XLSX_COLUMNS]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = BORDER

    for row_idx, record in enumerate(records, 2):
        is_alt = row_idx % 2 == 0
        for col_idx, key in enumerate(field_keys, 1):
            value = record.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = BORDER
            if is_alt:
                cell.fill = ALT_ROW_FILL

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for col_idx, key in enumerate(field_keys, 1):
        max_len = len(headers[col_idx - 1])
        for row in ws.iter_rows(min_row=2, max_row=min(len(records) + 1, 100), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        width = min(max_len + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(str(XLSX_PATH))
    print(f"Report generated: {XLSX_PATH} ({len(records)} rows)")
